from __future__ import annotations

import os
import tempfile
from io import BytesIO
from pathlib import Path

TEST_DB = Path(tempfile.gettempdir()) / "kpi_reset_import_test.db"
TEST_DB.unlink(missing_ok=True)
os.environ["DATABASE_URL"] = f"sqlite:///{TEST_DB.as_posix()}"
os.environ.setdefault("SECRET_KEY", "test-secret-test-secret-test-secret-123")
os.environ.setdefault("FRONTEND_URL", "http://localhost:5173")

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.seed import main as seed_main

seed_main()
from app.main import app

client = TestClient(app)


def login(email: str, password: str = "Admin@123"):
    response = client.post("/api/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


admin = login("admin@eaglesoftware.in")

# 1. Test Full System Factory Reset
reset_resp = client.post("/api/admin/reset-data", headers=admin, json={"confirm": "RESET", "mode": "full"})
assert reset_resp.status_code == 200, reset_resp.text
assert reset_resp.json()["ok"] is True

users_after_reset = client.get("/api/admin/users", headers=admin).json()
assert len(users_after_reset) == 1, f"Expected 1 user, got {len(users_after_reset)}"
assert users_after_reset[0]["email"] == "admin@eaglesoftware.in"

masters_after_reset = client.get("/api/admin/masters", headers=admin).json()
assert len(masters_after_reset) == 0 or all(len(d.get("departments", [])) == 0 for d in masters_after_reset)
assert client.get("/api/kpi/templates", headers=admin).json() == []
assert client.get("/api/kpi/cycles", headers=admin).json() == []

# 2. Build sample Excel file for import
wb = Workbook()
ws = wb.active
ws.append([
    "Employee No / Unique ID",
    "Full Name",
    "Email",
    "Temporary Password",
    "System Role",
    "Department",
    "Designation / Role",
    "Reporting Manager Email",
])
ws.append(["EMP-1001", "Sarah Jenkins", "sarah.jenkins@company.com", "Admin@123", "manager", "Operations", "Operations Manager", ""])
ws.append(["EMP-1002", "David Miller", "david.miller@company.com", "Admin@123", "employee", "Operations", "Operations Specialist", "sarah.jenkins@company.com"])
ws.append(["EMP-1003", "Anita Roy", "anita.roy@company.com", "Admin@123", "employee", "Quality Assurance", "QA Lead", "sarah.jenkins@company.com"])
buf = BytesIO()
wb.save(buf)
buf.seek(0)
excel_bytes = buf.getvalue()

# 3. Test Preview
preview_resp = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "true"},
    files={"file": ("employees.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert preview_resp.status_code == 200, preview_resp.text
pdata = preview_resp.json()
assert pdata["total_rows"] == 3
assert pdata["valid_rows"] == 3
assert pdata["created"] == 3
assert pdata["skipped"] == 0

# 4. Test actual import
import_resp = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "false"},
    files={"file": ("employees.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert import_resp.status_code == 200, import_resp.text
idata = import_resp.json()
assert idata["created"] == 3
assert idata["skipped"] == 0

users = client.get("/api/admin/users", headers=admin).json()
assert len(users) == 4  # Superadmin + 3 imported employees
sarah = next(u for u in users if u["email"] == "sarah.jenkins@company.com")
david = next(u for u in users if u["email"] == "david.miller@company.com")
assert sarah["employee_no"] == "EMP-1001"
assert david["employee_no"] == "EMP-1002"
assert david["manager_id"] == sarah["id"]

# 5. Test IDEMPOTENCY: Repeated import of the exact same workbook should skip without duplicate error
reimport_resp = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "false"},
    files={"file": ("employees.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert reimport_resp.status_code == 200, reimport_resp.text
rdata = reimport_resp.json()
assert rdata["created"] == 0
assert rdata["skipped"] == 3

# 6. Test CONFLICT: Same employee_no belonging to a different email must be blocked
wb_conflict = Workbook()
ws_c = wb_conflict.active
ws_c.append(["Employee No / Unique ID", "Full Name", "Email", "System Role", "Department", "Designation / Role"])
ws_c.append(["EMP-1001", "Fake Sarah", "fake.sarah@company.com", "employee", "Operations", "Operations Specialist"])
buf_c = BytesIO()
wb_conflict.save(buf_c)
buf_c.seek(0)

conflict_resp = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "false"},
    files={"file": ("conflict.xlsx", buf_c.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert conflict_resp.status_code == 400, "Conflicting employee ID must return 400"

# 7. Test DUPLICATE WITHIN FILE: Multiple rows with same employee_no must be blocked
wb_dup = Workbook()
ws_d = wb_dup.active
ws_d.append(["Employee No / Unique ID", "Full Name", "Email", "System Role", "Department", "Designation / Role"])
ws_d.append(["EMP-9999", "User One", "user1@company.com", "employee", "Operations", "Operations Specialist"])
ws_d.append(["EMP-9999", "User Two", "user2@company.com", "employee", "Operations", "Operations Specialist"])
buf_d = BytesIO()
wb_dup.save(buf_d)
buf_d.seek(0)

dup_resp = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "false"},
    files={"file": ("duplicate.xlsx", buf_d.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert dup_resp.status_code == 400, "Duplicate employee ID in file must return 400"

# 8. Test Second Reset & Re-import from scratch
reset_resp2 = client.post("/api/admin/reset-data", headers=admin, json={"confirm": "RESET", "mode": "full"})
assert reset_resp2.status_code == 200

users_cleared = client.get("/api/admin/users", headers=admin).json()
assert len(users_cleared) == 1
assert users_cleared[0]["email"] == "admin@eaglesoftware.in"

import_fresh = client.post(
    "/api/admin/import-employees-excel-v2",
    headers=admin,
    data={"preview": "false"},
    files={"file": ("employees.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
)
assert import_fresh.status_code == 200
assert import_fresh.json()["created"] == 3
users_restored = client.get("/api/admin/users", headers=admin).json()
assert len(users_restored) == 4

print("ALL RESET AND IDEMPOTENT IMPORT TESTS PASSED SUCCESSFULLY!")
