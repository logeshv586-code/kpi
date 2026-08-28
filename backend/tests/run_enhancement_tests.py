"""End-to-end regression checks for KPI System v1.2 enhancement features.

Run from the project root:
    DATABASE_URL=sqlite:////tmp/kpi_v12_test.db PYTHONPATH=backend python backend/tests/run_enhancement_tests.py
"""
from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

TEST_DB = Path("/tmp/kpi_v12_test.db")
TEST_DB.unlink(missing_ok=True)
os.environ.setdefault("DATABASE_URL", f"sqlite:///{TEST_DB}")
os.environ.setdefault("SECRET_KEY", "test-secret-test-secret-test-secret-123")
os.environ.setdefault("FRONTEND_URL", "http://localhost:5173")

from fastapi.testclient import TestClient
from openpyxl import Workbook
from reportlab.pdfgen import canvas

from app.seed import main as seed_main

seed_main()
from app.main import app

client = TestClient(app)


def login(email: str):
    response = client.post("/api/auth/login", json={"email": email, "password": "Admin@123"})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


admin = login("admin@eaglesoftware.in")
hr = login("hr@eaglesoftware.in")
assert client.post("/api/admin/reset-data", headers=hr, json={"confirm": "RESET"}).status_code == 403

rows = client.get("/api/kpi/my", headers=admin).json()
assignment = next(x for x in rows if x["status"] == "draft")
detail = client.get(f"/api/kpi/assignments/{assignment['id']}", headers=admin).json()
item = detail["template"]["kras"][0]["items"][0]

# XLSX KPI upload and assignment-aware question matching.
wb = Workbook(); ws = wb.active
ws.append(["KPI Parameter", "Actual Value", "Remarks", "Evidence File"])
ws.append([item["question"], 88, "Imported test", "proof.pdf"])
buf = BytesIO(); wb.save(buf); buf.seek(0)
upload = client.post("/api/files/upload", headers=admin, files={"file": ("kpi.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
assert upload.status_code == 200, upload.text
file_meta = upload.json()
parsed = client.post("/api/files/parse-kpi-excel", headers=admin, data={"file_id": file_meta["file_id"], "assignment_id": assignment["id"]})
assert parsed.status_code == 200, parsed.text
assert parsed.json()["matched"] >= 1
assert client.get(f"/api/files/{file_meta['file_id']}").status_code == 200
direct = client.post(f"/api/kpi/assignments/{assignment['id']}/import-responses", headers=admin, files={"file": ("kpi-direct.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
assert direct.status_code == 200, direct.text
assert direct.json()["matched"] >= 1

# Uploaded evidence reference persists on a KPI response.
payload = []
for kra in detail["template"]["kras"]:
    for kpi in kra["items"]:
        old = kpi.get("response") or {}
        row = {
            "kpi_item_id": kpi["id"],
            "actual_numeric": old.get("actual_numeric"),
            "answer_text": old.get("answer_text"),
            "selected_option": old.get("selected_option"),
            "remarks": old.get("remarks"),
            "evidence_url": old.get("evidence_url"),
            "evidence_file_id": old.get("evidence_file_id"),
        }
        if kpi["id"] == item["id"]:
            if kpi["input_type"] in {"choice", "yesno"}:
                row["selected_option"] = next(iter(kpi["config"]["score_map"]))
            else:
                row["actual_numeric"] = 88
            row["evidence_file_id"] = file_meta["file_id"]
        payload.append(row)
response = client.put(f"/api/kpi/assignments/{assignment['id']}/responses", headers=admin, json=payload)
assert response.status_code == 200, response.text
saved = client.get(f"/api/kpi/assignments/{assignment['id']}", headers=admin).json()
saved_item = next(i for k in saved["template"]["kras"] for i in k["items"] if i["id"] == item["id"])
assert saved_item["response"]["evidence_file_id"] == file_meta["file_id"]
assert saved_item["response"]["evidence_file"]["filename"] == "kpi.xlsx"

# PDF KPI import fallback parser.
pdf = BytesIO(); c = canvas.Canvas(pdf)
c.drawString(50, 800, "KPI Parameter | Actual Value | Remarks | Evidence File")
c.drawString(50, 780, f"{item['question']} | 91 | PDF imported | evidence.pdf")
c.save(); pdf.seek(0)
pdf_upload = client.post("/api/files/upload", headers=admin, files={"file": ("kpi.pdf", pdf.getvalue(), "application/pdf")})
assert pdf_upload.status_code == 200, pdf_upload.text
pdf_parse = client.post("/api/files/parse-kpi-excel", headers=admin, data={"file_id": pdf_upload.json()["file_id"], "assignment_id": assignment["id"]})
assert pdf_parse.status_code == 200, pdf_parse.text
assert len(pdf_parse.json()["rows"]) >= 1

# Excel template import.
wb = Workbook(); ws = wb.active
ws.append(["KRA", "KRA Weight", "KPI", "KPI Weight", "Input Type", "Target", "Direction", "Frequency", "Measurement", "Evidence Required"])
ws.append(["Delivery", 100, "On-time delivery", 100, "percentage", 100, "higher", "Monthly", "Percent on time", "Yes"])
buf = BytesIO(); wb.save(buf); buf.seek(0)
template = client.post("/api/kpi/templates/import-excel", headers=admin, data={"name": "Imported Test Template"}, files={"file": ("template.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
assert template.status_code == 200, template.text
assert template.json()["total_weight"] == 100
assert len(template.json()["kras"]) == 1
duplicate_import = client.post("/api/kpi/templates/import-excel", headers=admin, data={"name": "Imported Test Template"}, files={"file": ("template.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
assert duplicate_import.status_code == 409, duplicate_import.text

# Saving an edited draft must not keep its deleted KRA objects in the ORM
# collection.  Previously the second PUT validated 10 old + 10 new marks as
# 20 and returned HTTP 400; drafts are also intentionally allowed to be
# incomplete until publish.
draft_payload = {
    "name": "Incremental Draft Test",
    "division_id": None,
    "department_id": None,
    "designation_id": None,
    "kras": [{
        "name": "Delivery",
        "weight": 10,
        "items": [{
            "question": "Complete planned tasks",
            "input_type": "number",
            "weight": 10,
            "target_value": 10,
            "direction": "higher",
            "options": {},
        }],
    }],
}
draft_id = template.json()["id"]
# The normal lifecycle is publish -> unpublish -> edit draft -> publish again.
published = client.post(f"/api/kpi/templates/{draft_id}/publish", headers=admin)
assert published.status_code == 200, published.text
unpublished = client.post(f"/api/kpi/templates/{draft_id}/unpublish", headers=admin)
assert unpublished.status_code == 200, unpublished.text
assert unpublished.json()["status"] == "draft"
first_draft_save = client.put(f"/api/kpi/templates/{draft_id}", headers=admin, json=draft_payload)
assert first_draft_save.status_code == 200, first_draft_save.text
second_draft_save = client.put(f"/api/kpi/templates/{draft_id}", headers=admin, json=draft_payload)
assert second_draft_save.status_code == 200, second_draft_save.text
assert second_draft_save.json()["total_weight"] == 10
assert client.post(f"/api/kpi/templates/{draft_id}/publish", headers=admin).status_code == 400
complete_payload = {**draft_payload, "kras": [{
    **draft_payload["kras"][0], "weight": 100,
    "items": [{**draft_payload["kras"][0]["items"][0], "weight": 100}],
}]}
assert client.put(f"/api/kpi/templates/{draft_id}", headers=admin, json=complete_payload).status_code == 200
assert client.post(f"/api/kpi/templates/{draft_id}/publish", headers=admin).status_code == 200

# Employee workbook preview.
wb = Workbook(); ws = wb.active
ws.append(["Name", "Email", "Role", "Department", "Designation", "Manager Email"])
ws.append(["New Test User", "new.test@example.com", "employee", "Project Management", "Project Executive", "project@eaglesoftware.in"])
buf = BytesIO(); wb.save(buf); buf.seek(0)
employees = client.post("/api/admin/import-employees-excel", headers=admin, data={"preview": "true"}, files={"file": ("employees.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
assert employees.status_code == 200, employees.text
assert employees.json()["valid_rows"] == 1

# Super Admin can rename and remove an unused department from the employee form.
division_id = client.get("/api/admin/masters", headers=admin).json()[0]["id"]
department = client.post("/api/admin/departments", headers=admin, json={"name": "Rename Test Department", "parent_id": division_id})
assert department.status_code == 200, department.text
department_id = department.json()["id"]
renamed = client.put(f"/api/admin/departments/{department_id}", headers=admin, json={"name": "Renamed Test Department"})
assert renamed.status_code == 200, renamed.text
assert renamed.json()["name"] == "Renamed Test Department"
designation = client.post("/api/admin/designations", headers=admin, json={"name": "Rename Test Designation", "parent_id": department_id})
assert designation.status_code == 200, designation.text
designation_id = designation.json()["id"]
renamed_designation = client.put(f"/api/admin/designations/{designation_id}", headers=admin, json={"name": "Renamed Test Designation"})
assert renamed_designation.status_code == 200, renamed_designation.text
assert renamed_designation.json()["name"] == "Renamed Test Designation"
removed_designation = client.delete(f"/api/admin/designations/{designation_id}", headers=admin)
assert removed_designation.status_code == 200, removed_designation.text
removed_department = client.delete(f"/api/admin/departments/{department_id}", headers=admin)
assert removed_department.status_code == 200, removed_department.text

# KPI PDF summary export.
pdf_export = client.get(f"/api/kpi/assignments/{assignment['id']}/pdf", headers=admin)
assert pdf_export.status_code == 200
assert pdf_export.headers["content-type"].startswith("application/pdf")
assert len(pdf_export.content) > 1000

# Super Admin reset preserves masters/users/templates and removes transactional data.
before_users = len(client.get("/api/admin/users", headers=admin).json())
before_templates = len(client.get("/api/kpi/templates", headers=admin).json())
reset = client.post("/api/admin/reset-data", headers=admin, json={"confirm": "RESET"})
assert reset.status_code == 200, reset.text
assert client.get("/api/kpi/cycles", headers=admin).json() == []
assert client.get("/api/kpi/my", headers=admin).json() == []
assert len(client.get("/api/admin/users", headers=admin).json()) == before_users
assert len(client.get("/api/kpi/templates", headers=admin).json()) == before_templates

print("PASS: authentication and role restriction")
print("PASS: XLSX/PDF upload + parsing + KPI matching")
print("PASS: evidence_file_id persistence")
print("PASS: Excel template import")
print("PASS: publish, unpublish, repeated draft edit, and republish lifecycle")
print("PASS: employee Excel preview")
print("PASS: Super Admin department/designation rename and safe delete")
print("PASS: KPI PDF export")
print("PASS: Super Admin transactional data reset")
print("ALL KPI v1.2 ENHANCEMENT TESTS PASSED")
