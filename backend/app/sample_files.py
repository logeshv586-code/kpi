from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .file_storage import SAMPLE_DIR


def _save(name: str, headers: list[str], rows: list[list[object]]) -> Path:
    path = SAMPLE_DIR / name
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF2FF")
    for i, header in enumerate(headers, 1):
        width = max(len(str(header)) + 4, max((len(str(ws.cell(r, i).value or "")) for r in range(2, ws.max_row + 1)), default=0) + 2)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 52)
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def ensure_samples() -> dict[str, Path]:
    (SAMPLE_DIR / "KPI_Input_Sample.xlsx").unlink(missing_ok=True)

    return {
        "employees": _save(
            "Employee_Import_Sample.xlsx",
            [
                "Employee No / Unique ID",
                "Full Name",
                "Email",
                "Temporary Password",
                "System Role",
                "Department",
                "Designation / Role",
                "Reporting Manager Email",
            ],
            [
                [
                    "EMP-1001",
                    "Sarah Jenkins",
                    "sarah.jenkins@company.com",
                    "Admin@123",
                    "manager",
                    "Operations",
                    "Operations Manager",
                    "",
                ],
                [
                    "EMP-1002",
                    "David Miller",
                    "david.miller@company.com",
                    "Admin@123",
                    "employee",
                    "Operations",
                    "Operations Specialist",
                    "sarah.jenkins@company.com",
                ],
                [
                    "EMP-1003",
                    "Anita Roy",
                    "anita.roy@company.com",
                    "Admin@123",
                    "employee",
                    "Operations",
                    "Quality Lead",
                    "sarah.jenkins@company.com",
                ],
            ],
        ),
        "template": _save(
            "KPI_Template_Import_Sample.xlsx",
            [
                "KRA Name",
                "KRA Weight / Marks",
                "KPI Name",
                "Task Responsibility",
                "Result Entry Type",
                "Weight / Marks",
                "Expected Target",
                "Unit",
                "Scoring Direction",
                "Frequency",
                "Measurement / Guidance",
                "Custom Dropdown Results",
                "Source",
                "Weight Basis",
            ],
            [
                [
                    "Operational Excellence",
                    40,
                    "Process Tasks Completed",
                    "Complete all assigned operational tasks for the month",
                    "Number / Quantity",
                    20,
                    50,
                    "tasks",
                    "Higher result is better",
                    "Monthly",
                    "Enter total completed tasks count",
                    "",
                    "Task System",
                    "Configured by HR",
                ],
                [
                    "Operational Excellence",
                    40,
                    "SLA Compliance Rate",
                    "Maintain high SLA response and resolution percentage",
                    "Percentage",
                    20,
                    95,
                    "%",
                    "Higher result is better",
                    "Monthly",
                    "Enter actual SLA compliance percentage achieved",
                    "",
                    "Ticketing System",
                    "Configured by HR",
                ],
                [
                    "Quality & Client Feedback",
                    30,
                    "Client Audit Result",
                    "Client audit score and performance label selection",
                    "Custom Dropdown",
                    30,
                    "",
                    "",
                    "Higher result is better",
                    "Monthly",
                    "Select configured audit result from dropdown",
                    "Grade A - Exceeds Expectations=100; Grade B - Meets Target=80; Grade C - Satisfactory=60; Grade D - Needs Improvement=30",
                    "Audit Report",
                    "Configured by HR",
                ],
                [
                    "Team Leadership & Initiatives",
                    30,
                    "Training Completion Status",
                    "Complete mandatory team training and skill upgrades",
                    "Custom Dropdown",
                    30,
                    "",
                    "",
                    "Higher result is better",
                    "Monthly",
                    "Select training completion status",
                    "Completed All Modules=100; Partially Completed=50; Not Started=0",
                    "LMS Portal",
                    "Configured by HR",
                ],
            ],
        ),
    }
