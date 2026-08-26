from __future__ import annotations

import csv
import io
import os
import re
import shutil
import uuid
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
import pdfplumber

BASE_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = Path(os.getenv("KPI_UPLOAD_DIR", str(BASE_DIR / "uploads"))).resolve()
SAMPLE_DIR = Path(os.getenv("KPI_SAMPLE_DIR", str(BASE_DIR / "samples"))).resolve()
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
SAMPLE_DIR.mkdir(parents=True, exist_ok=True)

MAX_FILE_SIZE = int(os.getenv("KPI_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
UPLOAD_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv"}
TEMPLATE_EXTENSIONS = {".xlsx", ".xls", ".csv"}

CONTENT_TYPES = {
    ".pdf": "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".csv": "text/csv",
}


def safe_original_name(name: str | None) -> str:
    raw = Path(name or "upload").name
    raw = re.sub(r"[^A-Za-z0-9._() -]+", "_", raw).strip(" .")
    return raw[:180] or "upload"


def validate_extension(filename: str, allowed: set[str] | None = None) -> str:
    ext = Path(filename).suffix.lower()
    allowed = allowed or UPLOAD_EXTENSIONS
    if ext not in allowed:
        allowed_text = ", ".join(sorted(allowed))
        raise HTTPException(400, f"Unsupported file type. Allowed: {allowed_text}")
    return ext


async def save_upload(upload: UploadFile, allowed: set[str] | None = None) -> dict[str, Any]:
    original = safe_original_name(upload.filename)
    ext = validate_extension(original, allowed)
    file_id = uuid.uuid4().hex
    stored_name = f"{file_id}_{original}"
    target = UPLOAD_DIR / stored_name
    size = 0
    try:
        with target.open("wb") as out:
            while chunk := await upload.read(1024 * 1024):
                size += len(chunk)
                if size > MAX_FILE_SIZE:
                    out.close()
                    target.unlink(missing_ok=True)
                    raise HTTPException(413, f"File is larger than the {MAX_FILE_SIZE // (1024*1024)} MB limit")
                out.write(chunk)
    finally:
        await upload.close()
    return {
        "file_id": file_id,
        "filename": original,
        "stored_name": stored_name,
        "path": str(target),
        "url": f"/api/files/{file_id}",
        "content_type": upload.content_type or CONTENT_TYPES.get(ext, "application/octet-stream"),
        "size": size,
    }


def find_upload(file_id: str) -> Path | None:
    if not re.fullmatch(r"[0-9a-fA-F]{32}", str(file_id or "")):
        return None
    matches = list(UPLOAD_DIR.glob(f"{file_id}_*"))
    if not matches:
        return None
    path = matches[0].resolve()
    if UPLOAD_DIR not in path.parents:
        return None
    return path


def upload_metadata(file_id: str | None) -> dict[str, Any] | None:
    if not file_id:
        return None
    path = find_upload(file_id)
    if not path or not path.exists():
        return None
    prefix = f"{file_id}_"
    filename = path.name[len(prefix):] if path.name.startswith(prefix) else path.name
    ext = path.suffix.lower()
    return {
        "file_id": file_id,
        "filename": filename,
        "url": f"/api/files/{file_id}",
        "content_type": CONTENT_TYPES.get(ext, "application/octet-stream"),
        "size": path.stat().st_size,
    }


def clear_uploads() -> int:
    count = 0
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    for path in UPLOAD_DIR.iterdir():
        try:
            if path.is_file() or path.is_symlink():
                path.unlink()
                count += 1
            elif path.is_dir():
                shutil.rmtree(path)
                count += 1
        except FileNotFoundError:
            pass
    return count


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _rows_from_csv(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [{str(k or "").strip(): v for k, v in row.items()} for row in reader]


def _rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(x or "").strip() for x in next(rows)]
    except StopIteration:
        return []
    result = []
    for values in rows:
        if not any(v not in (None, "") for v in values):
            continue
        result.append({headers[i] if i < len(headers) else f"Column {i+1}": values[i] for i in range(len(values))})
    return result


def _rows_from_xls(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise HTTPException(500, "Legacy .xls support requires the xlrd package") from exc
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    if sheet.nrows < 1:
        return []
    headers = [str(sheet.cell_value(0, c) or "").strip() for c in range(sheet.ncols)]
    result = []
    for r in range(1, sheet.nrows):
        values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if not any(v not in (None, "") for v in values):
            continue
        result.append({headers[c]: values[c] for c in range(sheet.ncols)})
    return result


def _rows_from_pdf(path: Path) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table or len(table) < 2:
                    continue
                headers = [str(x or "").replace("\n", " ").strip() for x in table[0]]
                normalized = {_normalize_header(h) for h in headers}
                if not any("kpi" in h or "parameter" in h for h in normalized):
                    continue
                for values in table[1:]:
                    if not values or not any(v not in (None, "") for v in values):
                        continue
                    found.append({headers[i] if i < len(headers) else f"Column {i+1}": (values[i] if i < len(values) else None) for i in range(len(headers))})
        if found:
            return found
        # Fallback for simple pipe/tab separated PDF exports.
        lines: list[str] = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend([line.strip() for line in text.splitlines() if line.strip()])
        header_idx = next((i for i, line in enumerate(lines) if "kpi" in line.lower() and ("actual" in line.lower() or "value" in line.lower())), None)
        if header_idx is None:
            return []
        separator = "|" if "|" in lines[header_idx] else "\t" if "\t" in lines[header_idx] else None
        if not separator:
            return []
        headers = [x.strip() for x in lines[header_idx].split(separator)]
        for line in lines[header_idx + 1:]:
            if separator not in line:
                continue
            values = [x.strip() for x in line.split(separator)]
            found.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
    return found


def read_table(path: Path) -> list[dict[str, Any]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _rows_from_csv(path)
    if ext == ".xlsx":
        return _rows_from_xlsx(path)
    if ext == ".xls":
        return _rows_from_xls(path)
    if ext == ".pdf":
        return _rows_from_pdf(path)
    raise HTTPException(400, "This file type cannot be parsed")


def _get_value(row: dict[str, Any], *names: str) -> Any:
    lookup = {_normalize_header(k): v for k, v in row.items()}
    for name in names:
        key = _normalize_header(name)
        if key in lookup:
            return lookup[key]
    return None


def parse_response_rows(path: Path) -> list[dict[str, Any]]:
    rows = read_table(path)
    result = []
    for row in rows:
        parameter = _get_value(row, "KPI Parameter", "KPI", "Question", "Parameter", "Key Performance Indicator")
        actual = _get_value(row, "Actual Value", "Actual", "Achievement", "Answer", "Value")
        remarks = _get_value(row, "Remarks", "Remark", "Comments", "Comment")
        evidence = _get_value(row, "Evidence File", "Evidence", "Evidence URL", "Attachment")
        if parameter in (None, ""):
            continue
        result.append({
            "kpi_parameter": str(parameter).strip(),
            "actual_value": "" if actual is None else str(actual).strip(),
            "remarks": "" if remarks is None else str(remarks).strip(),
            "evidence_file": "" if evidence is None else str(evidence).strip(),
        })
    return result


def parse_template_rows(path: Path) -> list[dict[str, Any]]:
    rows = read_table(path)
    result = []
    for row in rows:
        kra = _get_value(row, "KRA", "Key Result Area", "Goal")
        kpi = _get_value(row, "KPI", "Key Performance Indicator", "Parameter", "KPI Parameter")
        if kra in (None, "") or kpi in (None, ""):
            continue
        result.append({
            "kra": str(kra).strip(),
            "kpi": str(kpi).strip(),
            "kra_weight": _get_value(row, "KRA Weight", "KRA Weightage", "KRA %"),
            "kpi_weight": _get_value(row, "KPI Weight", "Weight", "Weightage", "KPI %"),
            "measurement": _get_value(row, "Measurement", "Measure", "Instructions"),
            "target": _get_value(row, "Target", "Target Value"),
            "frequency": _get_value(row, "Frequency", "Periodicity"),
            "input_type": _get_value(row, "Input Type", "Type"),
            "direction": _get_value(row, "Direction", "Scoring Direction"),
            "evidence_required": _get_value(row, "Evidence Required", "Require Evidence"),
        })
    return result
